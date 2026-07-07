import openpyxl
import unicodedata
from datetime import date, datetime
from django.db import transaction
from django.http import HttpResponse

from academico.models import (
    Campus, Carrera, PeriodoDeNivelacion, Paralelo, EvaluacionAcademica,
    CohorteDeMatricula, InformeGeneral, ConsolidadoAcademico, MatriculaParalelo,
    Horario, MallaCurricular, UnidadCurricular)
from usuarios.models import PerfilEstudiante
from usuarios.utils import generar_identificador_siguiente

from poo.clases.enums.estado_de_matricula import EstadoDeMatricula
from poo.clases.enums.estado_de_aprobacion import EstadoDeAprobacion
from poo.clases.enums.estado_de_periodo import EstadoDePeriodo
from poo.clases.enums.modalidad import Modalidad
from poo.clases.enums.jornada import Jornada
from poo.clases.enums.tipo_de_cohorte import TipoDeCohorte
from poo.clases.enums.dia_de_semana import DiaDeSemana
from poo.clases.enums.estado_de_malla import EstadoDeMalla

from poo.clases.carrera import Carrera as CarreraBase
from poo.clases.periodo_de_nivelacion import PeriodoDeNivelacion as PeriodoDeNivelacionBase
from poo.clases.evaluacion_academica import EvaluacionAcademica as EvaluacionAcademicaPOO
from poo.clases.horario import Horario as HorarioPOO




# ══════════════════════════════════════════════════════════════
# UTILIDADES
# ══════════════════════════════════════════════════════════════

def normalizar_texto(texto):
    if not texto:
        return ""
    texto = str(texto).strip().lower()
    texto = ''.join(c for c in unicodedata.normalize('NFD', texto) if unicodedata.category(c) != 'Mn')
    return texto

def obtener_enum_flexible(enum_class, valor_sucio):
    if not valor_sucio:
        return None
    valor_normalizado = normalizar_texto(valor_sucio)
    for opcion in enum_class:
        if normalizar_texto(opcion.value) == valor_normalizado:
            return opcion
    raise ValueError(f"'{valor_sucio}' registro no válido para {enum_class.__name__}")



# ══════════════════════════════════════════════════════════════
# CARGA MASIVA: CAMPUS
# ══════════════════════════════════════════════════════════════

def servicio_campus_registrar_masivo_desde_excel(archivo, universidad_usuario):
    from poo.clases.campus import Campus as CampusBase

    resultado = {"exitosos": 0, "advertencias": [], "error": None}
    try:
        wb = openpyxl.load_workbook(archivo)
        ws = wb.active

        nombres_registrados = set()
        for nombre_existente in Campus.objects.filter(universidad=universidad_usuario).values_list("nombre", flat=True):
            nombres_registrados.add(normalizar_texto(nombre_existente))

        for numero_fila, fila in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            try:
                nombre = fila[0] if len(fila) > 0 else None
                direccion = fila[1] if len(fila) > 1 else None
                provincia = fila[2] if len(fila) > 2 else None

                if not nombre and not direccion and not provincia:
                    continue

                campus_poo = CampusBase(
                    codigo_de_campus="PENDIENTE",
                    nombre=str(nombre).strip() if nombre else "",
                    direccion_fisica=str(direccion).strip() if direccion else "",
                    provincia=str(provincia).strip() if provincia else "",
                )

                errores_validacion = campus_poo.validar_datos_de_carga_masiva()
                if errores_validacion:
                    resultado["advertencias"].append(
                        f"El registro de la fila {numero_fila} fue omitido por falta de información"
                    )
                    continue

                nombre_normalizado = normalizar_texto(campus_poo.nombre)
                if nombre_normalizado in nombres_registrados:
                    resultado["advertencias"].append(
                        f"El registro de la fila {numero_fila} fue omitido (el Campus ya existe)"
                    )
                    continue

                with transaction.atomic():
                    Campus.objects.create(
                        universidad=universidad_usuario,
                        codigo_de_campus=generar_identificador_siguiente(Campus, 'CAM', 'codigo_de_campus'),
                        nombre=campus_poo.nombre,
                        direccion_fisica=campus_poo.direccion_fisica,
                        provincia=campus_poo.provincia
                    )
                    resultado["exitosos"] += 1
                    nombres_registrados.add(nombre_normalizado)
            except Exception as e:
                resultado["advertencias"].append(f"El registro de la fila {numero_fila} fue omitido ({str(e)})")
                
    except Exception:
        resultado["error"] = "Ha ocurrido un error al procesar el documento"
        
    return resultado




# ══════════════════════════════════════════════════════════════
# CARGA MASIVA: CARRERAS
# ══════════════════════════════════════════════════════════════

def servicio_carrera_registrar_masivo_desde_excel(archivo, universidad_usuario):
    resultado = {"exitosos": 0, "advertencias": [], "error": None}
    try:
        wb = openpyxl.load_workbook(archivo)
        ws = wb.active
        
        campus_existente = Campus.objects.filter(universidad=universidad_usuario)

        carreras_registradas = {
            (campus_id, normalizar_texto(nombre_existente))
            for campus_id, nombre_existente in Carrera.objects.filter(
                campus__universidad=universidad_usuario
            ).values_list("campus_id", "nombre")
        }

        for numero_fila, fila in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            try:
                codigo_campus = fila[0] if len(fila) > 0 else None
                nombre = fila[1] if len(fila) > 1 else None
                vigencia = fila[2] if len(fila) > 2 else None
                
                if not codigo_campus and not nombre and not vigencia:
                    continue
                
                if not codigo_campus or not nombre or not vigencia:
                    resultado["advertencias"].append(f"El registro de la fila {numero_fila} fue omitido por falta de información")
                    continue

                if isinstance(vigencia, (datetime, date)):
                    vigencia_date = vigencia.date() if isinstance(vigencia, datetime) else vigencia
                elif isinstance(vigencia, str):
                    try:
                        vigencia_date = datetime.strptime(vigencia.strip(), "%Y-%m-%d").date()
                    except ValueError:
                        resultado["advertencias"].append(f"El registro de la fila {numero_fila} fue omitido (formato de fecha no válido)")
                        continue
                else:
                    resultado["advertencias"].append(f"El registro de la fila {numero_fila} fue omitido (formato de fecha no válido)")
                    continue

                campus_obj = campus_existente.filter(codigo_de_campus=str(codigo_campus).strip()).first()
                if not campus_obj:
                    resultado["advertencias"].append(f"El registro de la fila {numero_fila} fue omitido (código de Campus no válido)")
                    continue
                
                carrera_poo = CarreraBase(
                    codigo_de_carrera="PENDIENTE",
                    nombre=str(nombre).strip(),
                    vigencia_sniese=vigencia_date
                )

                if not carrera_poo.esta_activa():
                    resultado["advertencias"].append(f"El registro de la fila {numero_fila} fue omitido (Carrera no vigente)")
                    continue

                clave_carrera = (campus_obj.id, normalizar_texto(nombre))
                if clave_carrera in carreras_registradas:
                    resultado["advertencias"].append(
                        f"El registro de la fila {numero_fila} fue omitido (la Carrera ya existe)"
                    )
                    continue

                with transaction.atomic():
                    Carrera.objects.create(
                        campus=campus_obj,
                        codigo_de_carrera=generar_identificador_siguiente(Carrera, 'CAR', 'codigo_de_carrera'),
                        nombre=str(nombre).strip(),
                        vigencia_sniese=vigencia_date
                    )
                    resultado["exitosos"] += 1
                    carreras_registradas.add(clave_carrera)
            except Exception as e:
                resultado["advertencias"].append(f"El registro de la fila {numero_fila} fue omitido ({str(e)})")
                
    except Exception:
        resultado["error"] = "Ha ocurrido un error al procesar el documento"
        
    return resultado

# ══════════════════════════════════════════════════════════════
# CARGA MASIVA: MALLAS CURRICULARES
# ══════════════════════════════════════════════════════════════

def servicio_malla_registrar_masivo_desde_excel(archivo, universidad_usuario):
    from poo.clases.malla_curricular import MallaCurricular as MallaCurricularBase
    from poo.clases.enums.estado_de_malla import EstadoDeMalla

    resultado = {"exitosos": 0, "advertencias": [], "error": None}
    try:
        wb = openpyxl.load_workbook(archivo)
        ws = wb.active

        carreras_existentes = Carrera.objects.filter(campus__universidad=universidad_usuario)

        mallas_registradas = {
            (carrera_id, str(nombre_existente).strip().lower())
            for carrera_id, nombre_existente in MallaCurricular.objects.filter(
                carrera__campus__universidad=universidad_usuario
            ).values_list("carrera_id", "nombre")
        }

        for numero_fila, fila in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            try:
                codigo_carrera, nombre = fila[:2]

                if not any([codigo_carrera, nombre]):
                    continue

                if not all([codigo_carrera, nombre]):
                    resultado["advertencias"].append(
                        f"El registro de la fila {numero_fila} fue omitido por falta de información"
                    )
                    continue

                carrera_obj = carreras_existentes.filter(
                    codigo_de_carrera=str(codigo_carrera).strip()
                ).first()
                if not carrera_obj:
                    resultado["advertencias"].append(
                        f"El registro de la fila {numero_fila} fue omitido (código de Carrera no válido)"
                    )
                    continue

                malla_poo = MallaCurricularBase(
                    codigo_de_malla="PENDIENTE",
                    nombre=str(nombre).strip(),
                    version_de_malla="PENDIENTE",
                )

                errores_poo = malla_poo.validar_datos_de_registro()
                if errores_poo:
                    primer_error = list(errores_poo.values())[0]
                    resultado["advertencias"].append(
                        f"El registro de la fila {numero_fila} fue omitido ({primer_error})"
                    )
                    continue

                clave_malla = (carrera_obj.id, malla_poo.nombre.strip().lower())
                if clave_malla in mallas_registradas:
                    resultado["advertencias"].append(
                        f"El registro de la fila {numero_fila} fue omitido (la Malla curricular ya ha sido registrada)"
                    )
                    continue

                with transaction.atomic():
                    MallaCurricular.objects.create(
                        carrera=carrera_obj,
                        codigo_de_malla=generar_identificador_siguiente(
                            MallaCurricular, "MC", "codigo_de_malla"
                        ),
                        nombre=malla_poo.nombre,
                        version_de_malla=servicio_generar_version_malla(carrera_obj),
                        estado=EstadoDeMalla.DISENO.value,
                    )
                    resultado["exitosos"] += 1
                    mallas_registradas.add(clave_malla)

            except Exception as e:
                resultado["advertencias"].append(
                    f"El registro de la fila {numero_fila} fue omitido ({str(e)})"
                )

    except Exception:
        resultado["error"] = "Ha ocurrido un error al procesar el documento"

    return resultado



def servicio_unidad_registrar_masivo_desde_excel(archivo, universidad_usuario):
    from poo.clases.unidad_curricular import UnidadCurricular as UnidadCurricularBase
    from poo.clases.enums.estado_de_malla import EstadoDeMalla
    from academico.models import UnidadCurricular, MallaCurricular
    from usuarios.utils import generar_identificador_siguiente

    estados_editables = (EstadoDeMalla.DISENO.value, EstadoDeMalla.ACTIVA.value)

    resultado = {"exitosos": 0, "advertencias": [], "error": None}
    try:
        wb = openpyxl.load_workbook(archivo)
        ws = wb.active

        mallas_existentes = MallaCurricular.objects.filter(
            carrera__campus__universidad=universidad_usuario
        )

        unidades_registradas = {
            (malla_id, str(nombre_u).strip().lower())
            for malla_id, nombre_u in UnidadCurricular.objects.filter(
                malla_curricular__carrera__campus__universidad=universidad_usuario
            ).values_list("malla_curricular_id", "nombre")
        }

        for numero_fila, fila in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            try:
                (codigo_malla, nombre, horas_totales,
                 horas_sincronicas, horas_asincronicas,
                 criterio, porcentaje_asistencia) = fila[:7]

                if not any([codigo_malla, nombre, horas_totales,
                            horas_sincronicas, horas_asincronicas]):
                    continue

                if not all([codigo_malla, nombre, horas_totales,
                            horas_sincronicas, horas_asincronicas]):
                    resultado["advertencias"].append(
                        f"El registro de la fila {numero_fila} fue omitido por falta de información"
                    )
                    continue

                try:
                    horas_totales_f = float(horas_totales)
                    horas_sincronicas_f = float(horas_sincronicas)
                    horas_asincronicas_f = float(horas_asincronicas)
                    criterio_f = float(criterio) if criterio is not None else 7.0
                    porcentaje_f = float(porcentaje_asistencia) if porcentaje_asistencia is not None else 70.0
                except (ValueError, TypeError):
                    resultado["advertencias"].append(
                        f"El registro de la fila {numero_fila} fue omitido (registros numéricos no válidos)"
                    )
                    continue

                if horas_sincronicas_f < 6:
                    resultado["advertencias"].append(
                        f"El registro de la fila {numero_fila} fue omitido (mínimo de horas sincrónicas no válido (6))"
                    )
                    continue

                malla_obj = mallas_existentes.filter(
                    codigo_de_malla=str(codigo_malla).strip()
                ).first()
                if not malla_obj:
                    resultado["advertencias"].append(
                        f"El registro de la fila {numero_fila} fue omitido (código de Malla no válido)"
                    )
                    continue

                if malla_obj.estado not in estados_editables:
                    resultado["advertencias"].append(
                        f"El registro de la fila {numero_fila} fue omitido (estado no válido)"
                    )
                    continue

                clave_unidad = (malla_obj.id, str(nombre).strip().lower())
                if clave_unidad in unidades_registradas:
                    resultado["advertencias"].append(
                        f"El registro de la fila {numero_fila} fue omitido (la Unidad curricular ya ha sido registrada en la Malla especificada)"
                    )
                    continue

                unidad_poo = UnidadCurricularBase(
                    codigo_de_unidad="PENDIENTE",
                    nombre=str(nombre).strip(),
                    horas_totales=horas_totales_f,
                    horas_sincronicas=horas_sincronicas_f,
                    horas_asincronicas=horas_asincronicas_f,
                    criterio_de_aprobacion=criterio_f,
                    porcentaje_minimo_asistencia=porcentaje_f,
                )

                errores_poo = unidad_poo.validar_datos_de_registro()
                if errores_poo:
                    primer_error = list(errores_poo.values())[0]
                    resultado["advertencias"].append(
                        f"El registro de la fila {numero_fila} fue omitido ({primer_error})"
                    )
                    continue

                # Validar límite de horas sincrónicas semanales de la malla (20h)
                from poo.clases.franja_horaria import validar_malla_cabe_en_horario, SEMANAS_REFERENCIA_MINIMA
                from django.db.models import Sum

                suma_existente = UnidadCurricular.objects.filter(
                    malla_curricular=malla_obj
                ).aggregate(total=Sum("horas_sincronicas"))["total"] or 0.0
                total_con_nueva = suma_existente + horas_sincronicas_f
                validacion_malla = validar_malla_cabe_en_horario(total_con_nueva, SEMANAS_REFERENCIA_MINIMA)

                if not validacion_malla["ok"]:
                    resultado["advertencias"].append(
                        f"El registro de la fila {numero_fila} fue omitido (la Malla curricular excede el límite de {validacion_malla['limite']} horas sincrónicas semanales)"
                    )
                    continue

                with transaction.atomic():
                    UnidadCurricular.objects.create(
                        malla_curricular=malla_obj,
                        codigo_de_unidad=generar_identificador_siguiente(
                            UnidadCurricular, "UC", "codigo_de_unidad"
                        ),
                        nombre=unidad_poo.nombre,
                        horas_totales=unidad_poo.horas_totales,
                        horas_sincronicas=unidad_poo.horas_sincronicas,
                        horas_sincronicas_semanales=0,
                        horas_asincronicas=unidad_poo.horas_asincronicas,
                        criterio_de_aprobacion=unidad_poo.criterio_de_aprobacion,
                        porcentaje_minimo_asistencia=unidad_poo.porcentaje_minimo_asistencia,
                    )
                    resultado["exitosos"] += 1
                    unidades_registradas.add(clave_unidad)

            except Exception as e:
                resultado["advertencias"].append(
                    f"El registro de la fila {numero_fila} fue omitido ({str(e)})"
                )

    except Exception:
        resultado["error"] = "Ha ocurrido un error al procesar el documento"

    return resultado

def servicio_generar_paralelos(periodo_db, capacidad=35):
    import math
    from poo.clases.paralelo import Paralelo as ParaleloBase
    from poo.clases.servicios.centro_de_operacion_academica import CentroDeOperacionAcademica
    from poo.clases.cohorte_de_matricula import CohorteDeMatricula as CohorteDeMatriculaPOO
    from poo.clases.enums.jornada import Jornada
    from poo.clases.enums.modalidad import Modalidad as EnumModalidad
    from poo.clases.enums.estado_de_malla import EstadoDeMalla
    from poo.clases.enums.registro_de_cupo import RegistroDeCupo
    from poo.clases.enums.tipo_de_cohorte import TipoDeCohorte

    resumen = {
        "grupos_creados": 0,
        "paralelos_creados": 0,
        "estudiantes_distribuidos": 0,
        "advertencias": [],
    }

    try:
        capacidad = int(capacidad)
    except (TypeError, ValueError):
        capacidad = 35
    if capacidad <= 0:
        capacidad = 35

    universidad = periodo_db.universidad
    facade = CentroDeOperacionAcademica()
    enum_modalidad = obtener_enum_flexible(EnumModalidad, periodo_db.modalidad)

    import re

    def _num_sufijo(cadena):
        coincidencia = re.search(r"(\d+)$", str(cadena or ""))
        return int(coincidencia.group(1)) if coincidencia else 0

    contador_codigo = max(
        [_num_sufijo(c) for c in Paralelo.objects.values_list("codigo_de_paralelo", flat=True).distinct()] or [0],
        default=0,
    )

    carreras = Carrera.objects.filter(campus__universidad=universidad)

    for carrera in carreras:
        malla = MallaCurricular.objects.filter(
            carrera=carrera, estado=EstadoDeMalla.ACTIVA.value
        ).first()
        if not malla:
            if PerfilEstudiante.objects.filter(
                carrera_registrada=carrera,
                periodo_de_nivelacion=periodo_db,
                estado_de_matricula=EstadoDeMatricula.MATRICULADO.value,
            ).exists():
                resumen["advertencias"].append(
                    f"La Carrera {carrera.nombre} tiene estudiantes registrados pero no cuenta con una Malla curricular activa"
                )
            continue

        unidades = list(malla.unidades_curriculares.all())
        if not unidades:
            resumen["advertencias"].append(
                f"Los registros en la Carrera {carrera.nombre} fueron omitidos (sin registros asociados)"
            )
            continue

        jornadas_presentes = (
            PerfilEstudiante.objects.filter(
                carrera_registrada=carrera, periodo_de_nivelacion=periodo_db
            )
            .values_list("jornada", flat=True).distinct()
        )


        for jornada_valor in jornadas_presentes:
            estudiantes = list(
                PerfilEstudiante.objects.filter(
                    carrera_registrada=carrera,
                    jornada=jornada_valor,
                    periodo_de_nivelacion=periodo_db,
                    estado_de_matricula=EstadoDeMatricula.MATRICULADO.value,
                ).exclude(
                    estudiantes_matriculados__paralelo__periodo_de_nivelacion=periodo_db
                ).distinct()
            )
            if not estudiantes:
                continue

            try:
                enum_jornada = obtener_enum_flexible(Jornada, jornada_valor)
            except ValueError:
                resumen["advertencias"].append(
                    f"El registro de Jornada fue omitido (registro no válido)"
                )
                continue

            cohorte = _obtener_o_crear_cohorte(periodo_db, carrera)
            estudiantes_a_contar = []

            def _numero_de_grupo(nombre):
                try:
                    return int(str(nombre).split()[-1])
                except (ValueError, IndexError):
                    return 0

            with transaction.atomic():
                paralelos_existentes = Paralelo.objects.filter(
                    periodo_de_nivelacion=periodo_db,
                    jornada=jornada_valor,
                    unidad_curricular__in=unidades,
                )
                grupos_existentes = {}
                for paralelo_db in paralelos_existentes:
                    grupos_existentes.setdefault(paralelo_db.nombre, []).append(paralelo_db)

                indice_max = 0
                indice_pendiente = 0

                for nombre_grupo in sorted(grupos_existentes.keys(), key=_numero_de_grupo):
                    paralelos_grupo = grupos_existentes[nombre_grupo]
                    indice_max = max(indice_max, _numero_de_grupo(nombre_grupo))

                    representativo = paralelos_grupo[0]
                    ocupacion = MatriculaParalelo.objects.filter(paralelo=representativo).count()
                    cupo_libre = representativo.capacidad_maxima - ocupacion
                    if cupo_libre <= 0:
                        continue

                    a_matricular = estudiantes[indice_pendiente:indice_pendiente + cupo_libre]
                    indice_pendiente += len(a_matricular)

                    for paralelo_db in paralelos_grupo:
                        for estudiante_db in a_matricular:
                            MatriculaParalelo.objects.create(
                                estudiante=estudiante_db,
                                paralelo=paralelo_db,
                                cohorte_de_matricula=cohorte,
                            )
                    estudiantes_a_contar.extend(a_matricular)

                estudiantes_restantes = estudiantes[indice_pendiente:]
                if estudiantes_restantes:
                    # Calcular el siguiente índice de letra basado en el nombre más alto existente
                    # para evitar repetir códigos cuando se eliminan paralelos.
                    nombres_existentes = list(
                        Paralelo.objects.filter(
                            periodo_de_nivelacion=periodo_db,
                            unidad_curricular__malla_curricular__carrera=carrera,
                        ).values_list("nombre", flat=True).distinct()
                    )
                    # Encontrar el índice más alto de los nombres existentes
                    indice_base = 0
                    for nombre_existente in nombres_existentes:
                        # Parsear "Paralelo X" donde X es una letra como A, B, ..., Z, A1, B1...
                        nombre_limpio = nombre_existente.replace("Paralelo ", "").strip()
                        if len(nombre_limpio) == 1 and nombre_limpio.isalpha():
                            idx = ord(nombre_limpio.upper()) - ord('A') + 1
                        elif len(nombre_limpio) >= 2 and nombre_limpio[0].isalpha() and nombre_limpio[1:].isdigit():
                            idx = 26 + (int(nombre_limpio[1:]) - 1) * 26 + (ord(nombre_limpio[0].upper()) - ord('A')) + 1
                        else:
                            idx = 0
                        indice_base = max(indice_base, idx)

                    numero_de_grupos = math.ceil(len(estudiantes_restantes) / capacidad)
                    grupos_poo = [
                        ParaleloBase(
                            codigo_de_paralelo=f"G{indice}",
                            nombre=_nombre_paralelo_letra(indice_base + indice - 1),
                            jornada=enum_jornada,
                            modalidad=enum_modalidad,
                            capacidad_maxima=capacidad,
                        )
                        for indice in range(1, numero_de_grupos + 1)
                    ]

                    facade.distribuir_estudiantes(grupos_poo, estudiantes_restantes)

                    for indice, grupo_poo in enumerate(grupos_poo, start=1):
                        miembros = list(grupo_poo._estudiantes_matriculados)
                        if not miembros:
                            continue
                        nombre_nuevo = _nombre_paralelo_letra(indice_base + indice - 1)
                        contador_codigo += 1
                        codigo_nuevo = f"PAR{contador_codigo:03d}"
                        for unidad in unidades:
                            paralelo_db = Paralelo.objects.create(
                                periodo_de_nivelacion=periodo_db,
                                unidad_curricular=unidad,
                                codigo_de_paralelo=codigo_nuevo,
                                nombre=nombre_nuevo,
                                jornada=jornada_valor,
                                modalidad=periodo_db.modalidad,
                                capacidad_maxima=capacidad,
                            )
                            resumen["paralelos_creados"] += 1
                            for estudiante_db in miembros:
                                MatriculaParalelo.objects.create(
                                    estudiante=estudiante_db,
                                    paralelo=paralelo_db,
                                    cohorte_de_matricula=cohorte,
                                )
                        resumen["grupos_creados"] += 1
                        estudiantes_a_contar.extend(miembros)

                cohorte_poo = CohorteDeMatriculaPOO(
                    codigo_de_registro=cohorte.codigo_de_registro,
                    nombre_cohorte=cohorte.nombre_cohorte,
                    carrera_registrada=None,
                    fecha_de_cierre=periodo_db.fecha_fin,
                    periodo_de_nivelacion=None,
                    tipo_de_cohorte=obtener_enum_flexible(TipoDeCohorte, cohorte.tipo_de_cohorte),
                )
                for estudiante_db in estudiantes_a_contar:
                    cohorte_poo.registrar_estudiante_matriculado(estudiante_db)

                estadisticas_cohorte = cohorte_poo.obtener_estadisticas_de_registro()
                cohorte.total_primera_matricula += estadisticas_cohorte["Total primera matricula"]
                cohorte.total_segunda_matricula += estadisticas_cohorte["Total segunda matricula"]
                cohorte.total_exonerados += estadisticas_cohorte["Total exonerados"]
                resumen["estudiantes_distribuidos"] += len(estudiantes_a_contar)

                cohorte.save()

    return resumen

def servicio_mover_estudiante(estudiante_db, paralelo_destino_db):
    periodo = paralelo_destino_db.periodo_de_nivelacion
    carrera = paralelo_destino_db.unidad_curricular.malla_curricular.carrera
    nombre_destino = paralelo_destino_db.nombre
    jornada = paralelo_destino_db.jornada

    paralelos_destino = list(Paralelo.objects.filter(
        periodo_de_nivelacion=periodo,
        jornada=jornada,
        nombre=nombre_destino,
        unidad_curricular__malla_curricular__carrera=carrera,
    ))
    if not paralelos_destino:
        return (False, "La especificación del Paralelo de destino no es válida")

    representativo = paralelos_destino[0]
    ocupacion_destino = MatriculaParalelo.objects.filter(
        paralelo=representativo
    ).exclude(estudiante=estudiante_db).count()
    if ocupacion_destino >= representativo.capacidad_maxima:
        return (False, "El Paralelo de destino no presenta cupos disponibles")

    matriculas_actuales = MatriculaParalelo.objects.filter(
        estudiante=estudiante_db,
        paralelo__periodo_de_nivelacion=periodo,
        paralelo__unidad_curricular__malla_curricular__carrera=carrera,
    )

    primera_matricula = matriculas_actuales.first()
    if primera_matricula and primera_matricula.paralelo.nombre == nombre_destino:
        return (False, "El Estudiante ya pertenece al Paralelo especificado")

    cohorte = (
        primera_matricula.cohorte_de_matricula
        if primera_matricula else _obtener_o_crear_cohorte(periodo, carrera)
    )

    with transaction.atomic():
        matriculas_actuales.delete()
        for paralelo_db in paralelos_destino:
            MatriculaParalelo.objects.get_or_create(
                estudiante=estudiante_db,
                paralelo=paralelo_db,
                defaults={"cohorte_de_matricula": cohorte},
            )

    return (True, "El Estudiante fue reasignado correctamente")


def _paralelos_del_grupo_de_estudiantes(paralelo_db):
    carrera = paralelo_db.unidad_curricular.malla_curricular.carrera
    return Paralelo.objects.filter(
        periodo_de_nivelacion=paralelo_db.periodo_de_nivelacion,
        jornada=paralelo_db.jornada,
        nombre=paralelo_db.nombre,
        unidad_curricular__malla_curricular__carrera=carrera,
    )


def periodo_permite_gestion_matriculas(periodo_db):
    """Delega la consulta de estado al objeto POO PeriodoDeNivelacion."""
    periodo_poo = _construir_periodo(periodo_db)
    return periodo_poo.permite_gestion_matriculas()


def servicio_retirar_estudiante_de_paralelo(estudiante_db, paralelo_db):
    periodo = paralelo_db.periodo_de_nivelacion
    carrera = paralelo_db.unidad_curricular.malla_curricular.carrera

    if not periodo_permite_gestion_matriculas(periodo):
        return (False, "No se ha podido administrar la matrícula")

    paralelos_grupo = _paralelos_del_grupo_de_estudiantes(paralelo_db)
    matriculas = MatriculaParalelo.objects.filter(estudiante=estudiante_db, paralelo__in=paralelos_grupo)
    if not matriculas.exists():
        return (False, "El Estudiante no pertenece al Paralelo especificado")

    with transaction.atomic():
        matriculas.delete()
        estudiante_db.estado_de_matricula = EstadoDeMatricula.RETIRADO.value
        estudiante_db.save(update_fields=["estado_de_matricula"])

    servicio_recalcular_cohorte_de_carrera(periodo, carrera)
    return (True, "El Estudiante fue retirado del Paralelo correctamente")


def servicio_agregar_estudiante_a_paralelo(estudiante_db, paralelo_db):
    periodo = paralelo_db.periodo_de_nivelacion
    carrera = paralelo_db.unidad_curricular.malla_curricular.carrera

    if not periodo_permite_gestion_matriculas(periodo):
        return (False, "No se ha podido administrar la matrícula")

    if (estudiante_db.periodo_de_nivelacion_id != periodo.id
            or estudiante_db.carrera_registrada_id != carrera.id
            or estudiante_db.jornada != paralelo_db.jornada):
        return (False, "El Estudiante no es compatible con el Paralelo")

    if estudiante_db.estado_de_matricula in (EstadoDeMatricula.RETIRADO.value, EstadoDeMatricula.ANULADO.value):
        return (False, "El Estudiante no tiene una matrícula activa")

    if MatriculaParalelo.objects.filter(estudiante=estudiante_db, paralelo__periodo_de_nivelacion=periodo).exists():
        return (False, "El Estudiante ya se encuentra asignado a un Paralelo")

    paralelos_grupo = list(_paralelos_del_grupo_de_estudiantes(paralelo_db))
    representativo = paralelos_grupo[0]
    ocupacion = MatriculaParalelo.objects.filter(paralelo=representativo).count()
    if ocupacion >= representativo.capacidad_maxima:
        return (False, "El Paralelo no presenta cupos disponibles")

    cohorte = _obtener_o_crear_cohorte(periodo, carrera)
    with transaction.atomic():
        ya_matriculado = set(
            MatriculaParalelo.objects.filter(
                estudiante=estudiante_db, paralelo__in=paralelos_grupo
            ).values_list("paralelo_id", flat=True)
        )
        nuevas = [
            MatriculaParalelo(estudiante=estudiante_db, paralelo=p, cohorte_de_matricula=cohorte)
            for p in paralelos_grupo if p.id not in ya_matriculado
        ]
        if nuevas:
            MatriculaParalelo.objects.bulk_create(nuevas)

    servicio_recalcular_cohorte_de_carrera(periodo, carrera)
    return (True, "El Estudiante fue agregado al Paralelo correctamente")
